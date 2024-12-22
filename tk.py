import tkinter as tk
#from tkinter import ttk 
import ttkbootstrap as ttk
from tkinter import *
from PIL import ImageTk, Image
import openpyxl
import subprocess


button = None
button1 = None
button2 = None
button3 = None
back_button = None
message_label = None
title_label = None
treeview = None
back_button2 = None
back_button3 = None



def open_facture():
    subprocess.call(["python", "C:/Users/Hamza/Desktop/python tkinter/facture.py"])

def open_simulation():
    subprocess.call(["python", "C:/Users/Hamza/Desktop/python tkinter/credit.py"])

def excel_client(page):
    path = "C:/Users/Hamza/Desktop/python tkinter/tk.xlsx"
    workbook = openpyxl.load_workbook(path)
    if page == 1:
        sheet = workbook['client']

        list_values = list(sheet.values)
        #print(list_values)
        columns = list_values[0]  # Get the column names

        # Clear existing data in the treeview
        treeview.delete(*treeview.get_children())

        # Insert column headings into the treeview
        treeview['columns'] = columns
        for col in columns:
            treeview.heading(col, text=col)

        # Insert each row into the treeview
        for row in list_values[1:]:
            # Replace None values with empty strings
            row = ["" if cell is None else cell for cell in row]
            treeview.insert("", "end", values=row)
    if page == 2:

        sheet = workbook['catalogue']

        list_values = list(sheet.values)
        #print(list_values)
        columns = list_values[0]  # Get the column names

        # Clear existing data in the treeview
        treeview.delete(*treeview.get_children())

        # Insert column headings into the treeview
        treeview['columns'] = columns
        for col in columns:
            treeview.heading(col, text=col)

        # Insert each row into the treeview
        for row in list_values[1:]:
            # Replace None values with empty strings
            row = ["" if cell is None else cell for cell in row]
            treeview.insert("", "end", values=row)




window = ttk.Window(themename='darkly')
window.title('Facturation & Simulation de crédit')
window.geometry('955x650')



input_frame = ttk.Frame(master=window)
title_frame = ttk.Frame(master=input_frame, borderwidth=2, relief='ridge')
b_frame = ttk.Frame(master=input_frame)

def create_input_frame(status):
    global button, button1, button2, button3, title_label
    if status == 1:
        
        title_frame.pack(side='top', anchor='n', fill=None)

        b_frame.pack(side='right')

        button = ttk.Button(master=b_frame, text='Etablire une facture', command=open_facture)
        button1 = ttk.Button(master=b_frame, text='  Voire les clients    ', command=client)
        button2 = ttk.Button(master=b_frame, text=' Voire le catalogue ', command = voire_catalogue)
        button3 = ttk.Button(master=b_frame, text=' Simulation Credit  ', command =  open_simulation)

        button.pack(padx=100, pady=20)
        button1.pack(padx=100, pady=20)
        button2.pack(padx=100, pady=20)
        button3.pack(padx=100, pady=20)
        input_frame.pack(side='right', fill=BOTH, padx=20, pady=50)

        title_label = ttk.Label(master=title_frame, text='Facturation & Simulation de crédit', font='Calibri 13 bold',wraplength=200, justify='center', anchor='center')
        title_label.pack(padx=100, pady=15)
    else :
        if button:
            button.destroy()
        if button1:
            button1.destroy()
        if button2:
            button2.destroy()
        if button3:
            button3.destroy()
        if title_label:
            title_label.destroy()
        input_frame.pack_forget()
        title_frame.pack_forget()
        b_frame.pack_forget()

    

picture_frame = ttk.Frame(master=window)

def create_picture_frame(status):
    if status == 1 :
        picture_frame.pack(side='left')
        img = ImageTk.PhotoImage(Image.open("C:/Users/Hamza/Desktop/python tkinter/homepic.jpg"))

        display = Label(picture_frame, image=img)
        display.image = img  
        display.pack(fill=tk.BOTH, expand=True)
    else :
        picture_frame.pack_forget()
    

    

def home_page(status):
    create_input_frame(status)
    create_picture_frame(status)





new_frame = ttk.Frame(master=window)
treeframe = ttk.Frame(master=window)
bbtframe = ttk.Frame(master=window)
treescrool = ttk.Scrollbar(master=treeframe)
#treescrool.config(command=treeview)
cols = ('Numero', 'Raison sociale', 'Adresse')


def table(status, nsheet):
    global treeview
    if status == 1:
        treeframe.pack(side='right', fill= Y , expand= False)
        treeview = ttk.Treeview(master=treeframe, show='headings', columns=cols,yscrollcommand=treescrool.set, height=13)
        treeview.pack(side='right', fill= Y , expand= False, padx=20, pady=40)
        treescrool.pack(side="right", fill = Y)
        treescrool.config(command=treeview.set)
        excel_client(nsheet)     
    else :
        treeframe.destroy()
        treeview.pack_forget()


def client():
    global back_button, message_label, entry1, entry2, treeview, inser_bt
    home_page(0)
    new_frame.pack(side = 'left',fill=None, expand=False)
    # Back button
    bbtframe.pack(ipadx=30, ipady=10,anchor= tk.E)
    back_button = ttk.Button(master=bbtframe, text='Revenir', command=back_bt)
    back_button.pack(side='top')
    treescrool.pack(side="right", fill="y")
    
    # Message label
    message_label = ttk.Label(master=new_frame, text='Ajouter un client :')
    message_label.pack(side='top', padx=10, anchor = W)

    entry1 = ttk.Entry(new_frame)
    entry1.insert(0, "Entrer le nom")
    entry1.bind("<FocusIn>", lambda e: entry1.delete('0', 'end'))
    entry1.pack(padx=10, pady=10, anchor = W)

    entry2 = ttk.Entry(new_frame)
    entry2.insert(0, "Entrer l'adresse")
    entry2.bind("<FocusIn>", lambda b: entry2.delete('0', 'end'))
    entry2.pack(padx=10, pady=10, anchor = W)
    inser_bt= ttk.Button(master=new_frame, text='Ajouter', command=insert_client)
    inser_bt.pack(side='bottom')
    table(1, 1)




def back_bt():
    global back_button, message_label, entry1, entry2, treeview
    back_button.destroy()
    message_label.destroy()
    entry1.destroy()
    entry2.destroy()
    inser_bt.destroy()
    treeframe.pack_forget()
    new_frame.pack_forget()
    bbtframe.pack_forget()
    treescrool.pack_forget()
    if treeview :
        treeview.destroy()
    treescrool.pack_forget()
    home_page(1)


def insert_client():
    path = "C:/Users/Hamza/Desktop/python tkinter/tk.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['client']

    # Find the last row in the sheet
    last_row = sheet.max_row + 1

    # Get the input values from the Entry widgets
    raison_sociale = entry1.get()
    adresse = entry2.get()

    # Insert the values into the corresponding columns
    sheet.cell(row=last_row, column=1, value=last_row - 1)  # Numero
    sheet.cell(row=last_row, column=2, value=raison_sociale)  # Raison sociale
    sheet.cell(row=last_row, column=3, value=adresse)  # Adresse

    # Save the changes to the Excel file
    workbook.save(path)

    # Refresh the table view
    excel_client(1)


def voire_catalogue():
    global back_button2, message_label, entry1, entry2, treeview, inser_bt
    home_page(0)
    new_frame.pack(side = 'left',fill=None, expand=False)
    bbtframe.pack(ipadx=30, ipady=10,anchor= tk.E)
    back_button2 = ttk.Button(master=bbtframe, text='Revenir', command=back_bt2)
    back_button2.pack(side='top')
    treescrool.pack(side="right", fill="y")

    message_label = ttk.Label(master=new_frame, text='Ajouter un produit :')
    message_label.pack(side='top', padx=10, anchor = W)

    entry1 = ttk.Entry(new_frame)
    entry1.insert(0, "Entrer le nom")
    entry1.bind("<FocusIn>", lambda e: entry1.delete('0', 'end'))
    entry1.pack(padx=10, pady=10, anchor = W)

    entry2 = ttk.Entry(new_frame)
    entry2.insert(0, "Entrer le prix")
    entry2.bind("<FocusIn>", lambda b: entry2.delete('0', 'end'))
    entry2.pack(padx=10, pady=10, anchor = W)
    inser_bt= ttk.Button(master=new_frame, text='Ajouter', command=insert_produit)
    inser_bt.pack(side='bottom')
    table(1, 2)
    


def back_bt2():
    global back_button2, message_label, entry1, entry2, treeview

    back_button2.destroy()
    bbtframe.pack_forget()
    treescrool.pack_forget()
    new_frame.pack_forget()
    message_label.destroy()

    entry1.destroy()
    entry2.destroy()

    inser_bt.destroy()
    treeframe.pack_forget()
    if treeview :
        treeview.destroy()
    
    home_page(1)

def insert_produit():
    path = "C:/Users/Hamza/Desktop/python tkinter/tk.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['catalogue']

    # Find the last row in the sheet
    last_row = sheet.max_row + 1

    # Get the input values from the Entry widgets
    raison_sociale = entry1.get()
    adresse = entry2.get()

    # Insert the values into the corresponding columns
    sheet.cell(row=last_row, column=1, value=last_row - 1)  # Numero
    sheet.cell(row=last_row, column=2, value=raison_sociale)  # Raison sociale
    sheet.cell(row=last_row, column=3, value=adresse)  # Adresse

    # Save the changes to the Excel file
    workbook.save(path)

    # Refresh the table view
    excel_client(2)


def facture():
    global back_button3, message_label, entry1, entry2, treeview, inser_bt
    home_page(0)
    new_frame.pack(side = 'left',fill=None, expand=False)
    bbtframe.pack(ipadx=30, ipady=10,anchor= tk.E)
    back_button3 = ttk.Button(master=bbtframe, text='Revenir', command=back_bt3)
    back_button3.pack(side='top')


def back_bt3():
    global back_button3, message_label, entry1, entry2, treeview

    back_button3.destroy()
    bbtframe.pack_forget()
    treescrool.pack_forget()
    new_frame.pack_forget()

    home_page(1)

home_page(1)
#torun

window.mainloop()
